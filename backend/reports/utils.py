# reports/utils.py

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from django.http import HttpResponse
from datetime import datetime
from django.db import models

# --- NEW: SINGLE SOURCE OF TRUTH FOR HEADERS ---
def get_importable_headers(model_class):
    """
    Returns the definitive list of field names that should be included in
    an Excel template for importing. This is the single source of truth.
    """
    excluded_fields = ['id', 'user', 'department', 'created_at', 'updated_at']
    headers = []
    for field in model_class._meta.get_fields():
        # Include only concrete, non-relational fields that are not in the exclusion list
        if field.concrete and not field.is_relation and field.name not in excluded_fields:
            headers.append(field.name)
    return headers

def generate_excel_report(queryset, model_class):
    """
    Generates an Excel file from a given queryset and returns it as an HttpResponse.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    model_name = model_class._meta.verbose_name_plural.title()
    ws.title = model_name

    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    excluded_fields = ['id', 'user', 'department', 'created_at', 'updated_at']
    fields = [field for field in model_class._meta.get_fields() if field.name not in excluded_fields and not field.is_relation]
    
    headers = ['Faculty Name', 'Department'] + [field.verbose_name.title() for field in fields]

    ws.append(headers)
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for obj in queryset:
        row_data = [
            obj.faculty_name if hasattr(obj, 'faculty_name') else obj.user.get_full_name(),
            obj.department.name if obj.department else 'N/A'
        ]
        for field in fields:
            row_data.append(getattr(obj, field.name))
        ws.append(row_data)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = cell_alignment
            cell.border = thin_border

    for col_num, header_title in enumerate(headers, 1):
        max_length = 0
        column = get_column_letter(col_num)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M')
    filename = f"{model_name.replace(' ', '_')}_Report_{timestamp}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

def generate_blank_excel_template(model_class):
    wb = openpyxl.Workbook()
    
    instr_header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    instr_header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    instr_req_font = Font(name='Calibri', size=11, bold=True)
    required_highlight_fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')
    
    data_header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    data_header_fill = PatternFill(start_color='2F75B5', end_color='2F75B5', fill_type='solid')
    
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    red_font = Font(color='9C0006')
    dxf = DifferentialStyle(font=red_font, fill=red_fill)

    ws_instructions = wb.active
    ws_instructions.title = "Instructions"
    ws_instructions.append(["Column Name", "Required?", "Accepted Values / Format", "Description", "Example"])
    for cell in ws_instructions[1]:
        cell.font = instr_header_font
        cell.fill = instr_header_fill

    ws_data = wb.create_sheet(title="Data Entry")
    
    data_headers = get_importable_headers(model_class)
    field_map = {f.name: f for f in model_class._meta.get_fields()}

    ws_data.append(data_headers)

    for col_num, field_name in enumerate(data_headers, 1):
        field = field_map.get(field_name)
        if not field: continue

        required = "Yes" if not field.blank else "No"
        description = field.help_text or field.verbose_name.title()
        
        example, accepted_values = "", ""
        if field.choices:
            choices_list = [choice[0] for choice in field.choices]
            accepted_values, example = f"Choose from: {', '.join(choices_list)}", choices_list[0] if choices_list else ""
        elif isinstance(field, models.BooleanField):
            accepted_values, example = "Choose from: TRUE, FALSE", "TRUE"
        elif isinstance(field, models.DateField):
            accepted_values, example = "Format: YYYY-MM-DD", "2025-09-27"
        elif isinstance(field, models.URLField):
            accepted_values, example = "A valid URL (e.g., https://...)", "https://example.com/proof.pdf"
        elif isinstance(field, (models.IntegerField, models.DecimalField, models.PositiveIntegerField)):
            accepted_values, example = "Enter a number", "10"
        else:
            max_len = getattr(field, 'max_length', None)
            accepted_values, example = f"Text (max {max_len} chars)" if max_len else "Text", "Sample Text"
            
        instr_row = [field.name, required, accepted_values, description, example]
        ws_instructions.append(instr_row)
        if required == "Yes":
            for cell in ws_instructions[ws_instructions.max_row]:
                cell.font = instr_req_font
                cell.fill = required_highlight_fill

        header_cell = ws_data.cell(row=1, column=col_num)
        header_cell.value = field.name
        header_cell.font = data_header_font
        header_cell.fill = data_header_fill

        validation_formula = None
        if field.choices: validation_formula = f'"{",".join([c[0] for c in field.choices])}"'
        elif isinstance(field, models.BooleanField): validation_formula = '"TRUE,FALSE"'

        if validation_formula:
            dv = DataValidation(type="list", formula1=validation_formula, allow_blank=True)
            dv.error, dv.errorTitle = 'Your entry is not in the list.', 'Invalid Entry'
            ws_data.add_data_validation(dv)
            dv.add(f'{get_column_letter(col_num)}2:{get_column_letter(col_num)}1048576')
        
        if required == "Yes":
            ws_data.conditional_formatting.add(f'{get_column_letter(col_num)}2:{get_column_letter(col_num)}1048576', Rule(type="containsBlanks", dxf=dxf, stopIfTrue=True))

    for column_cells in ws_instructions.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws_instructions.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 5

    for i, column_title in enumerate(data_headers, 1):
        ws_data.column_dimensions[get_column_letter(i)].width = len(column_title) + 8
    
    ws_data.freeze_panes = 'A2'
    ws_instructions.protection.sheet = True
    wb.active = 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    model_name_str = model_class._meta.verbose_name.replace(' ', '_')
    filename = f"Template_{model_name_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response